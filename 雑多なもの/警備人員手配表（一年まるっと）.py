#CSVファイルを100個作るよ
from tkinter.tix import InputOnly
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import calendar
import datetime

Year= int(input('何年の分を作成しますか？：'))
for x in range(1,13):
    #日付の収集---------------------------------------------------------------------
    Now=datetime.datetime.now()
    Today= "{0:%Y/%m/%d}".format(Now)
    Month= x
    print('-'*40)
    #年と月から月終わりの日付を取得
    LastDay=(calendar.monthrange(Year,Month)[1])
    # date = datetime.date(Year,Month,1)#曜日を取得
    # Youbi=date.strftime('%a') 
    # print(f'{str(Year)}年{str(Month)}月分　警備手配表　{str(Today)}作成')
    # print('-'*40)
    #------------------------------------------------------------------------------
    #曜日を日本語で取得する関数
    def Youbi_get(Y0,M0,D0):
        date = datetime.date(Y0,M0,D0)
        NanYoubi0=date.strftime('%a')
        if NanYoubi0 == 'Sun':
            NanYoubi1= '日'
        elif NanYoubi0 == 'Mon':
            NanYoubi1= '月'
        elif NanYoubi0 == 'Tue':
            NanYoubi1= '火'
        elif NanYoubi0 == 'Wed':
            NanYoubi1= '水'
        elif NanYoubi0 == 'Thu':
            NanYoubi1= '木'
        elif NanYoubi0 == 'Fri':
            NanYoubi1= '金'
        else:
            NanYoubi1= '土'
        return NanYoubi1
    #エクセル操作-----------------------------------------------------------------
    path = r'C:\\Users\\ユーザー名\\Documents\\MyPython\\警備手配表\\'# ワークブックの読み込み
    wb = load_workbook(path+'警備手配表.xlsx')# ワークシートの選択
    ws = wb.active  # ワークシートを指定
    # セルに書き込み
    ws['A2'] = f'{str(Year)}年{str(Month)}月分　警備手配表　{str(Today)}更新'
    ws['A5'] = Month

    fill=openpyxl.styles.PatternFill(patternType='solid',fgColor='00FF99CC', bgColor='00FF99CC')#グレーd3d3d3
    for i in range(5,5+LastDay):
        ws['B'+str(i)]=i-4
        ws['C'+str(i)]=Youbi_get(Year,Month,i-4)
        if ws['C'+str(i)].value == '日':#日曜日の行を塗りつぶし
            for rows in ws['B'+str(i):'I'+str(i)]:
                for cell in rows:
                    cell.fill = fill

    # ワークブックに保存
    WBName=f'{str(Year)}年{str(Month)}月分　警備手配表.xlsx'
    wb.save(f'{path}{WBName}')
        
    print('完了-保存先：'+f'{path}{WBName}')