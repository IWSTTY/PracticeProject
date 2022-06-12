#CSVファイルを100個作るよ
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import calendar
import datetime

#日付の収集---------------------------------------------------------------------
Now=datetime.datetime.now()
Today= "{0:%Y/%m/%d}".format(Now)
Year= int("{0:%Y}".format(Now))
Month=int(input(f'（{str(Year)}年）月を入力して下さい：'))
print('-'*40)
#年と月から月終わりの日付を取得
LastDay=(calendar.monthrange(Year,Month)[1])
# date = datetime.date(Year,Month,1)#曜日を取得
# Youbi=date.strftime('%a') 
print(f'{str(Year)}年{str(Month)}月分　警備手配表　{str(Today)}更新')
print('-'*40)
#------------------------------------dd------------------------------------------
#曜日を日本語で取得する関数
def Youbi_get(Y0,M0,D0):
    date = datetime.date(Y0,M0,D0)
    NanYoubi0=date.strftime('%a')
    if NanYoubi0 == 'Sun':
        NanYoubi1= '日'
    elif NanYoubi0 == 'Mon':
        NanYoubi1= '月'
    elif NanYoubi0 == 'Tue':
        NanYoubi1= '火'
    elif NanYoubi0 == 'Wed':
        NanYoubi1= '水'
    elif NanYoubi0 == 'Thu':
        NanYoubi1= '木'
    elif NanYoubi0 == 'Fri':
        NanYoubi1= '金'
    else:
        NanYoubi1= '土'
    return NanYoubi1
#エクセル操作-----------------------------------------------------------------
path = r'C:\\Users\\ユーザー名\\Documents\\MyPython\\警備手配表\\'# ワークブックの読み込み
wb = load_workbook(path+'警備手配表.xlsx')# ワークシートの選択
ws = wb.active  # ワークシートを指定
# セルに書き込み
ws['A2'] = f'{str(Year)}年{str(Month)}月分　警備手配表　{str(Today)}更新'
ws['A5'] = Month

fill=openpyxl.styles.PatternFill(patternType='solid',fgColor='00FF99CC', bgColor='00FF99CC')#グレーd3d3d3
for i in range(5,5+LastDay):
    ws['B'+str(i)]=i-4
    ws['C'+str(i)]=Youbi_get(Year,Month,i-4)
    if ws['C'+str(i)].value == '日':#日曜日の行を塗りつぶし
        for rows in ws['B'+str(i):'I'+str(i)]:
            for cell in rows:
                cell.fill = fill

YesOrNo = input('常駐（月～土の昼間　1名）有り？【y or n】')
print('-'*40)
if YesOrNo == 'y':
    for i in range(5,5+LastDay):
        if ws['C'+str(i)].value == '日':
            continue
        else:
            ws['D'+str(i)]=1
else:
    print('白紙にて出力します')
    print('-'*40)

# ワークブックに保存
WBName=f'{str(Year)}年{str(Month)}月分　警備手配表.xlsx'
wb.save(f'{path}{WBName}')
    
print('完了-保存先：'+f'{path}{WBName}')